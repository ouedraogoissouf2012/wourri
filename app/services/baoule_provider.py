"""Provider Baoulé (#443) — validation JSON → Bronze uniquement.

N'écrit jamais pgvector. Pas de contenu baoulé inventé ici : on stocke
uniquement ce que le provider envoie après validation de forme.
"""
from __future__ import annotations

import json
import logging
from typing import Any

from app.data.lqe_languages import BAOULE_CODE
from app.services.improvement_queue import enqueue_improvement_task

logger = logging.getLogger(__name__)

REQUIRED = ("text_local", "text_fr")


def _as_entries(payload: Any) -> list[dict]:
    if isinstance(payload, list):
        return [x for x in payload if isinstance(x, dict)]
    if isinstance(payload, dict):
        if isinstance(payload.get("entries"), list):
            return [x for x in payload["entries"] if isinstance(x, dict)]
        return [payload]
    return []


def validate_baoule_entries(payload: Any) -> tuple[list[dict], list[str]]:
    """Retourne (entrées normalisées Bronze, erreurs)."""
    raw = _as_entries(payload)
    if not raw:
        return [], ["payload vide ou JSON invalide (attendu: liste d'objets)"]

    ok: list[dict] = []
    errors: list[str] = []
    for i, row in enumerate(raw):
        prefix = f"[{i}]"
        local = str(row.get("text_local") or "").strip()
        fr = str(row.get("text_fr") or "").strip()
        if not local:
            errors.append(f"{prefix} text_local requis")
            continue
        if not fr:
            errors.append(f"{prefix} text_fr requis")
            continue
        lang = str(row.get("language") or BAOULE_CODE).strip().lower()
        if lang not in {BAOULE_CODE, "baoule", "baoulé"}:
            errors.append(f"{prefix} language doit être {BAOULE_CODE} (reçu: {lang})")
            continue
        # status client ignoré — toujours bronze
        cultures = row.get("cultures")
        if cultures is not None and not isinstance(cultures, list):
            errors.append(f"{prefix} cultures doit être une liste")
            continue
        ok.append(
            {
                "external_id": str(row.get("id") or "").strip() or None,
                "language": BAOULE_CODE,
                "text_local": local[:2000],
                "text_fr": fr[:2000],
                "intent": str(row.get("intent") or "").strip() or None,
                "cultures": [str(c) for c in (cultures or [])][:20],
                "region": str(row.get("region") or "CI").strip()[:32],
                "notes": str(row.get("notes") or "").strip()[:500] or None,
                "source": "provider_upload",
                "status": "bronze",
            }
        )
    return ok, errors


def ingest_baoule_json(
    payload: Any,
    *,
    provider_id: str = "provider_baoule",
    path=None,
) -> dict:
    """Valide + écrit Bronze. Jamais de corpus prod."""
    entries, errors = validate_baoule_entries(payload)
    if not entries and errors:
        return {"ok": False, "accepted": 0, "rejected": 0, "errors": errors, "tasks": []}

    tasks = []
    skipped_dup = 0
    for ent in entries:
        excerpt = ent["text_local"]
        result = enqueue_improvement_task(
            intent=ent.get("intent"),
            source="provider_upload",
            cultures=ent.get("cultures") or [],
            excerpt=excerpt,
            user_anon=provider_id,
            path=path,
            language=BAOULE_CODE,
            skip_if_duplicate=True,
            extra={
                "text_fr": ent["text_fr"],
                "text_local": ent["text_local"],
                "external_id": ent.get("external_id"),
                "region": ent.get("region"),
                "notes": ent.get("notes"),
            },
        )
        if result.get("duplicate"):
            skipped_dup += 1
            continue
        if result.get("ok"):
            tasks.append(result.get("task"))
        else:
            errors.append(f"écriture refusée: {result.get('reason')}")

    return {
        "ok": len(tasks) > 0 or skipped_dup > 0,
        "accepted": len(tasks),
        "duplicates_skipped": skipped_dup,
        "rejected": len(errors),
        "errors": errors,
        "language": BAOULE_CODE,
        "tasks": tasks,
    }


def parse_json_bytes(data: bytes) -> Any:
    text = data.decode("utf-8-sig")
    return json.loads(text)


_HEADER_ALIASES = {
    "text_local": {
        "text_local",
        "local",
        "baoule",
        "baoule",
        "bci",
        "phrase_baoule",
        "phrase_locale",
        "texte_baoule",
        "texte_local",
        "source",
        "original",
        "texte",
        "phrase",
        "a_traduire",
        "atraduire",
    },
    "text_fr": {
        "text_fr",
        "fr",
        "francais",
        "french",
        "phrase_fr",
        "traduction",
        "texte_fr",
        "texte_francais",
        "target",
        "cible",
        "translation",
    },
    "id": {"id", "identifiant", "ref", "n", "no", "numero"},
    "intent": {"intent", "intention"},
    "cultures": {"cultures", "culture"},
    "region": {"region"},
    "notes": {"notes", "note", "commentaire"},
    "language": {"language", "langue", "lang"},
}


def _fold(s: str) -> str:
    """Minuscule + sans accents + espaces → _ pour matcher les en-têtes Excel."""
    import unicodedata

    t = unicodedata.normalize("NFD", str(s or ""))
    t = "".join(c for c in t if unicodedata.category(c) != "Mn")
    t = t.strip().lower().replace(" ", "_").replace("-", "_")
    while "__" in t:
        t = t.replace("__", "_")
    return t


def _norm_header(h: str) -> str | None:
    key = _fold(h)
    for canon, aliases in _HEADER_ALIASES.items():
        folded_aliases = {_fold(a) for a in aliases}
        if key in folded_aliases or key == canon:
            return canon
    return None


def _rows_from_table(headers: list[str], rows: list[list[Any]]) -> list[dict]:
    mapped = [_norm_header(h) for h in headers]
    if "text_local" not in mapped and "text_fr" not in mapped:
        # 2 colonnes sans en-tête reconnu → col0=local, col1=fr
        if len(headers) >= 2 and not any(mapped):
            mapped = ["text_local", "text_fr"] + [None] * max(0, len(headers) - 2)
    out: list[dict] = []
    for row in rows:
        item: dict[str, Any] = {}
        for i, canon in enumerate(mapped):
            if not canon or i >= len(row):
                continue
            val = row[i]
            if val is None:
                continue
            s = str(val).strip()
            if not s:
                continue
            if canon == "cultures":
                item[canon] = [c.strip() for c in s.replace(";", ",").split(",") if c.strip()]
            else:
                item[canon] = s
        if item.get("text_local") or item.get("text_fr"):
            out.append(item)
    return out


def _decode_text(data: bytes) -> str:
    """UTF-8 puis Windows Excel (cp1252) puis latin-1."""
    for enc in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            return data.decode(enc)
        except UnicodeDecodeError:
            continue
    return data.decode("utf-8", errors="replace")


def parse_csv_bytes(data: bytes) -> list[dict]:
    import csv
    import io

    text = _decode_text(data)
    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
    except csv.Error:
        dialect = csv.get_dialect("excel")
        # Excel FR exporte souvent en ;
        if sample.count(";") >= sample.count(","):
            class _Semi(csv.excel):
                delimiter = ";"
            dialect = _Semi
    reader = csv.reader(io.StringIO(text), dialect)
    rows = list(reader)
    if not rows:
        raise ValueError("CSV vide")
    headers, body = rows[0], rows[1:]
    result = _rows_from_table(headers, body)
    if not result:
        raise ValueError(
            "Aucune ligne lue. En-têtes trouvés: "
            + ", ".join(repr(h) for h in headers[:12])
            + " — utilise text_local (ou baoule) et text_fr (ou francais)."
        )
    return result


def parse_xlsx_bytes(data: bytes) -> list[dict]:
    import io

    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError(
            "openpyxl non installé sur le serveur — Redeploy l'image API"
        ) from exc
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    last_headers: list[str] = []
    for ws in wb.worksheets:
        rows_iter = ws.iter_rows(values_only=True)
        try:
            headers = [str(c) if c is not None else "" for c in next(rows_iter)]
        except StopIteration:
            continue
        last_headers = headers
        body = [list(r) for r in rows_iter]
        result = _rows_from_table(headers, body)
        if result:
            return result
    raise ValueError(
        "Excel: aucune ligne valide. En-têtes: "
        + ", ".join(repr(h) for h in last_headers[:12])
        + " — colonnes requises: text_local (baoule) + text_fr (francais)."
    )


def parse_upload(filename: str, data: bytes) -> Any:
    """JSON / CSV / XLSX → liste d'objets pour validate_baoule_entries."""
    name = (filename or "").lower().strip()
    if name.endswith(".json"):
        return parse_json_bytes(data)
    if name.endswith(".csv") or name.endswith(".txt"):
        return parse_csv_bytes(data)
    if name.endswith(".xlsx") or name.endswith(".xlsm"):
        return parse_xlsx_bytes(data)
    # signature xlsx = ZIP PK
    if len(data) >= 2 and data[0:2] == b"PK":
        return parse_xlsx_bytes(data)
    head = data.lstrip()[:1]
    if head in (b"[", b"{"):
        return parse_json_bytes(data)
    # défaut: tenter CSV (Excel FR)
    try:
        return parse_csv_bytes(data)
    except Exception:
        pass
    raise ValueError("Format non supporté — utilise .json, .csv ou .xlsx")
