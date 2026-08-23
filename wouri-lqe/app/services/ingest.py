"""Ingest générique — status toujours bronze, langue du compte, backend Postgres."""
from __future__ import annotations

import csv
import io
import json
import logging
import unicodedata
from typing import Any

from app.data.language_registry import is_known
from app.services import productions_repo as repo
from app.services.fingerprint import fingerprint

logger = logging.getLogger(__name__)

HEADER_ALIASES = {
    "text_local": {"text_local", "local", "source", "original", "texte", "phrase"},
    "text_fr": {"text_fr", "fr", "francais", "french", "traduction", "cible"},
    "id": {"id", "identifiant", "ref"},
    "concept_id": {"concept_id", "concept", "id_concept"},
    "intent": {"intent", "intention"},
    "cultures": {"cultures", "culture"},
    "region": {"region"},
    "notes": {"notes", "note"},
    "audio_url": {"audio_url", "audio", "audio_ref"},
}


def _fold(s: str) -> str:
    t = unicodedata.normalize("NFD", str(s or ""))
    t = "".join(c for c in t if unicodedata.category(c) != "Mn")
    return t.strip().lower().replace(" ", "_").replace("-", "_")


def _norm_header(h: str) -> str | None:
    key = _fold(h)
    for canon, aliases in HEADER_ALIASES.items():
        if key == canon or key in {_fold(a) for a in aliases}:
            return canon
    return None


def _rows_from_table(headers: list[str], rows: list[list[Any]]) -> list[dict]:
    mapped = [_norm_header(h) for h in headers]
    if "text_local" not in mapped and "text_fr" not in mapped and len(headers) >= 2 and not any(mapped):
        mapped = ["text_local", "text_fr"] + [None] * max(0, len(headers) - 2)
    out = []
    for row in rows:
        item: dict[str, Any] = {}
        for i, canon in enumerate(mapped):
            if not canon or i >= len(row) or row[i] is None:
                continue
            s = str(row[i]).strip()
            if not s:
                continue
            if canon == "cultures":
                item[canon] = [c.strip() for c in s.replace(";", ",").split(",") if c.strip()]
            else:
                item[canon] = s
        if item.get("text_local") or item.get("text_fr"):
            out.append(item)
    return out


def _decode(data: bytes) -> str:
    for enc in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            return data.decode(enc)
        except UnicodeDecodeError:
            continue
    return data.decode("utf-8", errors="replace")


def parse_upload(filename: str, data: bytes) -> list[dict]:
    name = (filename or "").lower()
    if name.endswith(".json") or data.lstrip()[:1] in (b"[", b"{"):
        payload = json.loads(_decode(data))
        if isinstance(payload, dict) and isinstance(payload.get("entries"), list):
            return [x for x in payload["entries"] if isinstance(x, dict)]
        if isinstance(payload, list):
            return [x for x in payload if isinstance(x, dict)]
        if isinstance(payload, dict):
            return [payload]
        return []
    if name.endswith(".xlsx") or name.endswith(".xlsm") or data[:2] == b"PK":
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        ws = wb.active
        it = ws.iter_rows(values_only=True)
        headers = [str(c) if c is not None else "" for c in next(it)]
        body = [list(r) for r in it]
        return _rows_from_table(headers, body)
    text = _decode(data)
    try:
        dialect = csv.Sniffer().sniff(text[:4096], delimiters=",;\t")
    except csv.Error:
        dialect = csv.excel
        if text.count(";") >= text.count(","):
            dialect = type("S", (csv.excel,), {"delimiter": ";"})()
    rows = list(csv.reader(io.StringIO(text), dialect))
    if not rows:
        return []
    return _rows_from_table(rows[0], rows[1:])


def ingest(entries: list[dict], *, language: str, actor: str) -> dict:
    if not is_known(language):
        return {"ok": False, "accepted": 0, "duplicates_skipped": 0, "errors": ["langue inconnue"]}
    accepted = 0
    skipped = 0
    errors = []
    for i, raw in enumerate(entries):
        local = str(raw.get("text_local") or "").strip()
        fr = str(raw.get("text_fr") or "").strip()
        if not local or not fr:
            errors.append(f"[{i}] text_local et text_fr requis")
            continue
        ext = str(raw.get("id") or raw.get("external_id") or "").strip() or None
        fp = fingerprint(language=language, text_local=local, text_fr=fr, external_id=ext)
        concept_id = str(raw.get("concept_id") or "").strip() or None
        meta = {
            "region": str(raw.get("region") or "CI"),
            "notes": str(raw.get("notes") or ""),
            "source": "provider_upload",
            "external_id": ext,
            "actor": actor,
        }
        try:
            res = repo.insert_bronze(
                language=language,
                text_local=local[:2000],
                text_fr=fr[:2000],
                intent=str(raw.get("intent") or ""),
                cultures=raw.get("cultures") if isinstance(raw.get("cultures"), list) else [],
                audio_url=raw.get("audio_url") or raw.get("audio_ref"),
                fingerprint=fp,
                concept_id=concept_id,
                created_by=actor,
                meta=meta,
            )
        except Exception as exc:  # une ligne fautive n'annule pas le lot (tolerance par ligne)
            logger.warning("ingest ligne %s: %s", i, exc)
            errors.append(f"[{i}] insert")
            continue
        if res["inserted"]:
            accepted += 1
        else:
            skipped += 1
    return {
        "ok": accepted > 0 or skipped > 0,
        "accepted": accepted,
        "duplicates_skipped": skipped,
        "errors": errors,
        "language": language,
    }
