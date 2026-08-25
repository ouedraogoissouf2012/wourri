"""Dictee guidee — logique metier de la collecte du dataset ASR (ADR-0035).

Deux responsabilites, sans SQL (delegue a `dictation_repo`) :
  1. `parse_prompts` : lit un lot de phrases (CSV / XLSX / JSON) -> [{filiere, text_fr, text_local}].
  2. `build_export_zip` : empaquete le dataset au format HF `audiofolder`
     (dossier `audio/` + `metadata.csv` : file_name, transcription, language, filiere, text_fr).

Le texte est IMPOSE au locuteur (transcription = `text_local`, garantie) ; ce module ne
touche pas au flux de parite (productions).
"""
from __future__ import annotations

import csv
import io
import json
import unicodedata
import zipfile
from pathlib import Path
from typing import Any

from app.services import dictation_repo as repo
from app.services.audio_store import AudioStore, LocalAudioStore

# Colonnes acceptees a l'import (accents/casse/espaces normalises par `_fold`).
_HEADER_ALIASES = {
    "filiere": {"filiere", "filiere_agricole", "secteur", "culture"},
    "text_fr": {"text_fr", "francais", "french", "fr", "question", "question_fr", "traduction"},
    "text_local": {"text_local", "baoule", "bci", "local", "reponse", "phrase", "cible"},
}


def _fold(s: str) -> str:
    t = unicodedata.normalize("NFD", str(s or ""))
    t = "".join(c for c in t if unicodedata.category(c) != "Mn")
    return t.strip().lower().replace(" ", "_").replace("-", "_")


def _norm_header(h: str) -> str | None:
    key = _fold(h)
    for canon, aliases in _HEADER_ALIASES.items():
        if key == canon or key in {_fold(a) for a in aliases}:
            return canon
    return None


def _decode(data: bytes) -> str:
    for enc in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            return data.decode(enc)
        except UnicodeDecodeError:
            continue
    return data.decode("utf-8", errors="replace")


def _rows_from_table(headers: list[str], rows: list[list[Any]]) -> list[dict]:
    mapped = [_norm_header(h) for h in headers]
    # Aucune colonne reconnue : on suppose l'ordre du lot locuteur (filiere, fr, baoule),
    # ou (fr, baoule) a deux colonnes. Une phrase locale (text_local) reste obligatoire.
    if not any(mapped):
        if len(headers) >= 3:
            mapped = ["filiere", "text_fr", "text_local"] + [None] * (len(headers) - 3)
        elif len(headers) == 2:
            mapped = ["text_fr", "text_local"]
    out: list[dict] = []
    for row in rows:
        item: dict[str, str] = {}
        for i, canon in enumerate(mapped):
            if not canon or i >= len(row) or row[i] is None:
                continue
            s = str(row[i]).strip()
            if s:
                item[canon] = s
        if item.get("text_local"):
            out.append({
                "filiere": item.get("filiere", ""),
                "text_fr": item.get("text_fr", ""),
                "text_local": item["text_local"],
            })
    return out


def _rows_from_json(payload: Any) -> list[dict]:
    items = payload.get("prompts") if isinstance(payload, dict) else payload
    out: list[dict] = []
    for x in items or []:
        if not isinstance(x, dict):
            continue
        local = str(x.get("text_local") or x.get("baoule") or "").strip()
        if not local:
            continue
        out.append({
            "filiere": str(x.get("filiere") or "").strip(),
            "text_fr": str(x.get("text_fr") or x.get("francais") or "").strip(),
            "text_local": local,
        })
    return out


def parse_prompts(filename: str, data: bytes) -> list[dict]:
    """Lit un lot de phrases (CSV/XLSX/JSON) et normalise en [{filiere, text_fr, text_local}].
    Une ligne sans `text_local` (la phrase a lire) est ignoree."""
    name = (filename or "").lower()
    if name.endswith(".json") or data.lstrip()[:1] in (b"[", b"{"):
        return _rows_from_json(json.loads(_decode(data)))
    if name.endswith((".xlsx", ".xlsm")) or data[:2] == b"PK":
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        ws = wb.active
        it = ws.iter_rows(values_only=True)
        headers = [str(c) if c is not None else "" for c in next(it)]
        body = [list(r) for r in it]
        return _rows_from_table(headers, body)
    text = _decode(data)
    try:
        dialect: Any = csv.Sniffer().sniff(text[:4096], delimiters=",;\t")
    except csv.Error:
        dialect = csv.excel
        if text.count(";") > text.count(","):
            dialect = type("S", (csv.excel,), {"delimiter": ";"})()
    rows = list(csv.reader(io.StringIO(text), dialect))
    if not rows:
        return []
    return _rows_from_table(rows[0], rows[1:])


def build_export_zip(*, language: str, store: AudioStore | None = None) -> tuple[bytes, int]:
    """Empaquete le dataset ASR de `language` au format HF `audiofolder` :
    `audio/<fichier>` + `metadata.csv` (file_name, transcription, language, filiere, text_fr).
    `transcription` = le texte IMPOSE (baoule). Retourne (octets_zip, nb_clips). Les prompts
    dont le fichier audio est introuvable sont ignores (jamais d'entree orpheline)."""
    audio = store or LocalAudioStore()
    rows = repo.export_rows(language=language)
    buf = io.BytesIO()
    n = 0
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        meta = io.StringIO()
        writer = csv.writer(meta)
        writer.writerow(["file_name", "transcription", "language", "filiere", "text_fr"])
        for r in rows:
            ref = r.get("audio_url")
            if not ref:
                continue
            try:
                clip = audio.load(ref)
            except FileNotFoundError:
                continue
            arc = f"audio/{Path(ref).name}"
            zf.writestr(arc, clip)
            writer.writerow([arc, r.get("text_local", ""), language,
                             r.get("filiere", ""), r.get("text_fr", "")])
            n += 1
        zf.writestr("metadata.csv", meta.getvalue())
    return buf.getvalue(), n
